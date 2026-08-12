#!/usr/bin/env python3
"""
Bulk import/export of question content between a spreadsheet and questions.json.

WHY THIS IS NOT A BACKEND
Authoring content in bulk, validating it, and scheduling it are all build-time
concerns — the player's browser never needs to talk to a server for any of them. A
database would add cost, latency, an attack surface, and a second place for content
to live, and would solve none of the problems below. Content stays in git, so every
change is reviewable in a pull request and revertible, which is worth a great deal
for a game whose credibility rests on every number being checkable.

WHY VALIDATION IS THE POINT
A bulk import is the fastest possible way to ship wrong numbers. Schema validation
here checks the SHAPE — required fields, types, ranges, unique ids, domains that
actually contain their points. It does not and cannot check whether a stat is true.
That is verify_questions.py's job, and it re-derives every number from the source
datasets. Both gates have to pass:

    edit spreadsheet
      -> import_questions.py     shape is right
      -> verify_questions.py     numbers are real
      -> build_site.py           ship it

ROUND TRIP
Export first, edit, import back. The export is the authoring surface; questions.json
is the build artifact.

    python3 pipeline/import_questions.py --export content/questions.xlsx
    python3 pipeline/import_questions.py --export content/questions.csv
    python3 pipeline/import_questions.py --import content/questions.xlsx
    python3 pipeline/import_questions.py --import content/questions.csv --dry-run

SCHEDULING
An optional schedule pins specific questions to specific dates. Anything unpinned
falls back to the existing deterministic shuffled bag, so the schedule is an
override, not a replacement — and the daily selection stays a pure function of the
date with no server involved.

    content/schedule.csv:  date,question_id      (one row per question per date)
    python3 pipeline/import_questions.py --schedule content/schedule.csv
"""
import argparse, csv, json, os, re, sys
from collections import Counter, defaultdict

HERE = os.path.dirname(os.path.abspath(__file__))
DATA = os.path.join(HERE, '..', 'data')
QJSON = os.path.join(DATA, 'questions.json')
SCHEDULE_JSON = os.path.join(DATA, 'schedule.json')
SCHEMA = os.path.join(DATA, 'questions.schema.json')
MAX_REFS = 5

BASE_COLS = ['id', 'league',
             'x_label', 'x_unit', 'x_step', 'x_min', 'x_max',
             'y_label', 'y_unit', 'y_step', 'y_min', 'y_max',
             'target_player', 'target_x', 'target_y']
REF_COLS = [f'ref{i}_{f}' for i in range(1, MAX_REFS + 1) for f in ('name', 'x', 'y')]
TAIL_COLS = ['fact', 'source']
COLUMNS = BASE_COLS + REF_COLS + TAIL_COLS


# ---------------------------------------------------------------- flatten / inflate
def to_row(q):
    row = {
        'id': q['id'], 'league': q['league'],
        'x_label': q['xLabel'], 'x_unit': q['xUnit'], 'x_step': q.get('xStep', ''),
        'x_min': q['xDomain'][0], 'x_max': q['xDomain'][1],
        'y_label': q['yLabel'], 'y_unit': q['yUnit'], 'y_step': q.get('yStep', ''),
        'y_min': q['yDomain'][0], 'y_max': q['yDomain'][1],
        'target_player': q['targetPlayer'],
        'target_x': q['targetX'], 'target_y': q['targetY'],
        'fact': q['fact'], 'source': q['source'],
    }
    for i, r in enumerate(q['referencePlayers'][:MAX_REFS], start=1):
        row[f'ref{i}_name'], row[f'ref{i}_x'], row[f'ref{i}_y'] = r['name'], r['x'], r['y']
    for c in COLUMNS:
        row.setdefault(c, '')
    return row


def num(v, field, rid, errors, required=True):
    if v in ('', None):
        if required:
            errors.append(f'{rid}: {field} is required and empty')
        return None
    try:
        f = float(v)
    except (TypeError, ValueError):
        errors.append(f'{rid}: {field} is not a number ({v!r})')
        return None
    return int(f) if f == int(f) else f


def to_question(row, errors):
    rid = (row.get('id') or '').strip() or '<no id>'
    q = {
        'id': rid,
        'league': (row.get('league') or '').strip().upper(),
        'xLabel': (row.get('x_label') or '').strip(),
        'xUnit': (row.get('x_unit') or '').strip(),
        'yLabel': (row.get('y_label') or '').strip(),
        'yUnit': (row.get('y_unit') or '').strip(),
        'xDomain': [num(row.get('x_min'), 'x_min', rid, errors),
                    num(row.get('x_max'), 'x_max', rid, errors)],
        'yDomain': [num(row.get('y_min'), 'y_min', rid, errors),
                    num(row.get('y_max'), 'y_max', rid, errors)],
        'targetPlayer': (row.get('target_player') or '').strip(),
        'targetX': num(row.get('target_x'), 'target_x', rid, errors),
        'targetY': num(row.get('target_y'), 'target_y', rid, errors),
        'referencePlayers': [],
        'fact': (row.get('fact') or '').strip(),
        'source': (row.get('source') or '').strip(),
    }
    for k, col in (('xStep', 'x_step'), ('yStep', 'y_step')):
        v = num(row.get(col), col, rid, errors, required=False)
        if v is not None:
            q[k] = v
    for i in range(1, MAX_REFS + 1):
        name = (row.get(f'ref{i}_name') or '').strip()
        if not name:
            continue
        q['referencePlayers'].append({
            'name': name,
            'x': num(row.get(f'ref{i}_x'), f'ref{i}_x', rid, errors),
            'y': num(row.get(f'ref{i}_y'), f'ref{i}_y', rid, errors),
        })
    # Field order matches the hand-authored questions so diffs stay readable.
    ordered = ['id', 'league', 'xLabel', 'xUnit', 'yLabel', 'yUnit', 'xStep', 'yStep',
               'xDomain', 'yDomain', 'targetPlayer', 'targetX', 'targetY',
               'referencePlayers', 'fact', 'source']
    return {k: q[k] for k in ordered if k in q}


# ---------------------------------------------------------------- validation
def validate(questions):
    """Structural validation against questions.schema.json, plus the cross-field rules
    a JSON Schema can't express. Returns a list of human-readable errors."""
    errors = []
    schema = json.load(open(SCHEMA))
    props = schema['$defs']['question']['properties']
    required = schema['$defs']['question']['required']
    leagues = props['league']['enum']
    id_pat = re.compile(props['id']['pattern'])

    seen_ids = Counter(q.get('id', '') for q in questions)
    for dup, n in seen_ids.items():
        if n > 1:
            errors.append(f'{dup}: duplicate id appears {n} times')

    for q in questions:
        rid = q.get('id') or '<no id>'
        for field in required:
            if field not in q or q[field] in ('', None, []):
                errors.append(f'{rid}: missing required field {field}')
        if not id_pat.match(rid):
            errors.append(f'{rid}: id must be kebab-case (a-z, 0-9, hyphens)')
        if q.get('league') not in leagues:
            errors.append(f'{rid}: league must be one of {leagues}, got {q.get("league")!r}')

        for txt, lo, hi in (('fact', props['fact']['minLength'], props['fact']['maxLength']),
                            ('source', props['source']['minLength'], props['source']['maxLength'])):
            v = q.get(txt) or ''
            if v and not (lo <= len(v) <= hi):
                errors.append(f'{rid}: {txt} is {len(v)} chars, must be {lo}-{hi}')

        for axis in ('x', 'y'):
            dom = q.get(f'{axis}Domain') or []
            step = q.get(f'{axis}Step', 0.1)
            if len(dom) != 2 or None in dom:
                errors.append(f'{rid}: {axis}Domain must be two numbers')
                continue
            if dom[0] >= dom[1]:
                errors.append(f'{rid}: {axis}Domain min must be below max ({dom})')
            if step is not None and step <= 0:
                errors.append(f'{rid}: {axis}Step must be positive, got {step}')
            # Every plotted point must be inside the plotted range, or it draws
            # off-chart. The game widens silently to cope; content shouldn't rely on it.
            pts = [(q.get(f'target{axis.upper()}'), q.get('targetPlayer'))]
            pts += [(r.get(axis), r.get('name')) for r in q.get('referencePlayers', [])]
            for val, who in pts:
                if val is None:
                    continue
                if not (dom[0] <= val <= dom[1]):
                    errors.append(f'{rid}: {who} {axis}={val} is outside {axis}Domain {dom}')

        refs = q.get('referencePlayers') or []
        if not (2 <= len(refs) <= 5):
            errors.append(f'{rid}: needs 2-5 reference players, has {len(refs)}')
        # Distinct people, ignoring season suffixes — two seasons of one player as two
        # separate dots reads as a bug, not as a comparison.
        base = [re.sub(r',?\s*\d{4}(-\d{2})?$', '', r.get('name', '')).strip().lower()
                for r in refs]
        for nm, n in Counter(base).items():
            if n > 1:
                errors.append(f'{rid}: "{nm}" appears as {n} separate reference dots')
        tgt = re.sub(r',?\s*\d{4}(-\d{2})?$', '', q.get('targetPlayer', '')).strip().lower()
        if tgt in base:
            errors.append(f'{rid}: target "{tgt}" is also a reference player')
    return errors


# ---------------------------------------------------------------- spreadsheet io
def read_table(path):
    if path.lower().endswith(('.xlsx', '.xlsm')):
        try:
            import openpyxl
        except ImportError:
            sys.exit('reading .xlsx needs openpyxl:  pip install openpyxl\n'
                     '(or export/import .csv instead — same columns, no dependency)')
        wb = openpyxl.load_workbook(path, data_only=True)
        ws = wb['questions'] if 'questions' in wb.sheetnames else wb.active
        rows = list(ws.iter_rows(values_only=True))
        header = [str(c).strip() if c is not None else '' for c in rows[0]]
        return [dict(zip(header, ['' if c is None else c for c in r]))
                for r in rows[1:] if any(c not in ('', None) for c in r)]
    with open(path, newline='', encoding='utf-8-sig') as fh:
        return [r for r in csv.DictReader(fh) if any(v.strip() for v in r.values() if v)]


def write_table(path, rows):
    os.makedirs(os.path.dirname(os.path.abspath(path)) or '.', exist_ok=True)
    if path.lower().endswith('.xlsx'):
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment
        except ImportError:
            sys.exit('writing .xlsx needs openpyxl:  pip install openpyxl')
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'questions'
        ws.append(COLUMNS)
        for c in ws[1]:
            c.font = Font(bold=True)
        for r in rows:
            ws.append([r.get(c, '') for c in COLUMNS])
        for i, col in enumerate(COLUMNS, start=1):
            letter = openpyxl.utils.get_column_letter(i)
            ws.column_dimensions[letter].width = (
                60 if col in ('fact', 'source') else
                26 if col.endswith('_label') or col.endswith('player') or col.endswith('_name')
                else 12)
        for row in ws.iter_rows(min_row=2):
            for c in row:
                c.alignment = Alignment(vertical='top', wrap_text=True)
        ws.freeze_panes = 'A2'
        wb.save(path)
        return
    with open(path, 'w', newline='', encoding='utf-8') as fh:
        w = csv.DictWriter(fh, fieldnames=COLUMNS, extrasaction='ignore')
        w.writeheader()
        w.writerows(rows)


# ---------------------------------------------------------------- schedule
def build_schedule(path):
    """date,question_id rows -> {date: [ids]}. Pinned dates override the shuffled
    bag; everything else keeps falling back to it, so this stays a pure function of
    the date and needs no server."""
    by_date = defaultdict(list)
    for r in read_table(path):
        date = str(r.get('date', '')).strip()[:10]
        qid = str(r.get('question_id', '')).strip()
        if not date or not qid:
            continue
        if not re.match(r'^\d{4}-\d{2}-\d{2}$', date):
            sys.exit(f'schedule: bad date {date!r} (want YYYY-MM-DD)')
        by_date[date].append(qid)

    known = {q['id'] for q in json.load(open(QJSON))}
    problems = []
    for date, ids in sorted(by_date.items()):
        for qid in ids:
            if qid not in known:
                problems.append(f'{date}: unknown question id {qid!r}')
        dupes = [i for i, n in Counter(ids).items() if n > 1]
        if dupes:
            problems.append(f'{date}: same question pinned twice: {dupes}')
        if len(ids) != 5:
            problems.append(f'{date}: pins {len(ids)} questions, a day serves 5')
    if problems:
        print('SCHEDULE ERRORS:')
        for p in problems:
            print('  -', p)
        return 1
    json.dump(dict(sorted(by_date.items())), open(SCHEDULE_JSON, 'w'), indent=2)
    open(SCHEDULE_JSON, 'a').write('\n')
    print(f'wrote {os.path.relpath(SCHEDULE_JSON)}: {len(by_date)} scheduled days')
    return 0


# ---------------------------------------------------------------- main
def main():
    ap = argparse.ArgumentParser()
    g = ap.add_mutually_exclusive_group(required=True)
    g.add_argument('--export', metavar='PATH', help='questions.json -> .xlsx/.csv')
    g.add_argument('--import', dest='imp', metavar='PATH', help='.xlsx/.csv -> questions.json')
    g.add_argument('--schedule', metavar='PATH', help='date,question_id csv -> schedule.json')
    g.add_argument('--check', action='store_true', help='validate questions.json as-is')
    ap.add_argument('--dry-run', action='store_true', help='validate an import without writing')
    a = ap.parse_args()

    if a.schedule:
        return build_schedule(a.schedule)

    if a.check:
        errors = validate(json.load(open(QJSON)))
        print(f'{len(json.load(open(QJSON)))} questions, {len(errors)} problems')
        for e in errors:
            print('  -', e)
        return 1 if errors else 0

    if a.export:
        questions = json.load(open(QJSON))
        write_table(a.export, [to_row(q) for q in questions])
        print(f'exported {len(questions)} questions -> {a.export}')
        print('Edit it, then:  python3 pipeline/import_questions.py --import ' + a.export)
        return 0

    rows = read_table(a.imp)
    errors = []
    questions = [to_question(r, errors) for r in rows]
    errors += validate(questions)
    print(f'read {len(questions)} rows from {a.imp}; {len(errors)} problems')
    for e in errors:
        print('  -', e)
    if errors:
        print('\nREFUSED — nothing written. Fix the rows above and re-run.')
        return 1
    if a.dry_run:
        print('\ndry run — nothing written. Looks importable.')
        return 0
    json.dump(questions, open(QJSON, 'w'), indent=2, ensure_ascii=False)
    open(QJSON, 'a').write('\n')
    print(f'\nwrote {len(questions)} questions -> {os.path.relpath(QJSON)}')
    print('NEXT: python3 pipeline/verify_questions.py   '
          '(schema validity is not truth — the numbers still need re-deriving)')
    return 0


if __name__ == '__main__':
    sys.exit(main())
